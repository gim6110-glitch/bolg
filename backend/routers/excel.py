from __future__ import annotations

import json
from io import BytesIO
from typing import Any

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from services import generate_record, mask_personal_info, review_record, safe_area

router = APIRouter(prefix="/api/excel", tags=["excel"])
AREAS = "자율활동,동아리활동,진로활동,과목별 세부능력 및 특기사항,행동특성 및 종합의견"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WARNING_FILL = PatternFill("solid", fgColor="FCE4E4")


def style_header(ws, row: int, columns: int) -> None:
    for cell in ws[row][:columns]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def build_template_workbook() -> Workbook:
    wb = Workbook()
    guide = wb.active
    guide.title = "사용안내"
    guide.merge_cells("A1:E1")
    guide["A1"] = "개인정보(이름, 학번, 주민등록번호, 연락처)는 입력하지 마세요. 발견 시 마스킹됩니다."
    guide["A1"].fill = WARNING_FILL
    guide["A1"].font = Font(color="C00000", bold=True)
    guide.append(["시트", "용도"])
    guide.append(["작성용", "교사 관찰·활동 내용을 입력하여 작성 결과를 받습니다."])
    guide.append(["검토용", "기존 작성 문구를 입력하여 위반사항과 수정안을 받습니다."])
    style_header(guide, 2, 2)

    write = wb.create_sheet("작성용")
    write.append(["개인정보 입력 금지: 이름·학번·연락처 등은 자동 마스킹됩니다."])
    write.merge_cells("A1:E1")
    write["A1"].fill = WARNING_FILL
    write["A1"].font = Font(color="C00000", bold=True)
    write.append(["연번", "식별자", "영역", "과목명(세특만)", "관찰·활동 내용"])
    style_header(write, 2, 5)
    write.freeze_panes = "A3"

    review = wb.create_sheet("검토용")
    review.append(["개인정보 입력 금지: 이름·학번·연락처 등은 자동 마스킹됩니다."])
    review.merge_cells("A1:E1")
    review["A1"].fill = WARNING_FILL
    review["A1"].font = Font(color="C00000", bold=True)
    review.append(["연번", "식별자", "영역", "기존 작성 문구", "글자수"])
    style_header(review, 2, 5)
    review.freeze_panes = "A3"
    for row in range(3, 103):
        review[f"E{row}"] = f'=IF(D{row}="","",LEN(D{row}))'

    dv = DataValidation(type="list", formula1=f'"{AREAS}"', allow_blank=False)
    for ws in (write, review):
        ws.add_data_validation(dv)
        dv.add(f"C3:C102")
        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 24 if col != "E" else 60
    return wb


@router.get("/template")
def template() -> StreamingResponse:
    wb = build_template_workbook()
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=saenggibu_template.xlsx"},
    )


def append_result(ws, row_data: list[Any]) -> None:
    ws.append(row_data)


@router.post("/upload")
def upload(file: UploadFile = File(...)) -> StreamingResponse:
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="xlsx 파일만 업로드할 수 있습니다.")
    wb = load_workbook(file.file)
    if "결과" in wb.sheetnames:
        del wb["결과"]
    result_ws = wb.create_sheet("결과")
    result_ws.append(["모드", "연번", "식별자", "영역", "원문/입력", "결과 JSON", "개인정보경고"])
    style_header(result_ws, 1, 7)

    if "작성용" in wb.sheetnames:
        for row in wb["작성용"].iter_rows(min_row=3, values_only=True):
            number, identifier, area, subject, content = row[:5]
            if not area or not content:
                continue
            area = safe_area(str(area))
            masked, detected = mask_personal_info(str(content))
            payload = {"과목명": subject or "", "관찰·활동 내용": masked}
            response = generate_record(area, payload)
            append_result(result_ws, ["작성", number, identifier, area, content, json.dumps(response, ensure_ascii=False), "감지" if detected else ""])

    if "검토용" in wb.sheetnames:
        for row in wb["검토용"].iter_rows(min_row=3, values_only=True):
            number, identifier, area, text = row[:4]
            if not area or not text:
                continue
            area = safe_area(str(area))
            masked, detected = mask_personal_info(str(text))
            response = review_record(area, masked)
            append_result(result_ws, ["검토", number, identifier, area, text, json.dumps(response, ensure_ascii=False), "감지" if detected else ""])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=saenggibu_results.xlsx"},
    )
