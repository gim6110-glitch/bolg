import os
import sys

import pytest

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
BACKEND = os.path.join(ROOT, 'backend')
if BACKEND not in sys.path:
    sys.path.insert(0, BACKEND)

from prompts.builder import build_review_prompt, build_write_prompt, get_max_chars, normalize_area


def test_prompt_builder_is_2015_curriculum_and_area_specific():
    system, user = build_write_prompt('세특', {'과목명': '기계일반', '과제': '도면 해석'})

    assert normalize_area('세특') == '과목별 세부능력 및 특기사항'
    assert '2015 개정 교육과정' in system
    assert '2022 개정 영역명으로 바꾸지 않음' in system
    assert '공백 포함 300~350자. (잘한 학생은 최대한 350자, 부족하면 억지로 늘리지 않음)' in system
    assert "성취수준은 '상 수준임', '도달함' 등 직접 판정 표현 금지" in system
    assert '"과목명": "기계일반"' in user
    assert get_max_chars('행발') == 300


def test_review_prompt_contains_checklist_and_schema():
    system, user = build_review_prompt('진로활동', '진로탐색대회에서 수상하였다.')

    assert '검토 체크리스트' in system
    assert '대학명·기관명·기업명·강사명' in system
    assert '기존 작성 문구' in user
    assert '위반사항' in user
    assert '원문위치' in user
    assert '종합의견' in user
    assert '최대글자수: 500' in user


def test_personal_info_masking_and_similarity_when_dependencies_available():
    pytest.importorskip('anthropic')
    from services import find_local_similarities, mask_personal_info

    masked, detected = mask_personal_info('홍길동 12번 010-1234-5678')

    assert detected is True
    assert '[개인정보 제거]' in masked

    result = find_local_similarities([
        {'식별자': '1', '문장': '자료를 비교 분석하며 문제 해결 과정을 구체화함.'},
        {'식별자': '2', '문장': '자료를 비교 분석하며 문제 해결 과정을 정리함.'},
    ])
    assert result['이상없음'] is False
    assert result['유사쌍'][0]['유사구절']


def test_excel_template_structure_when_dependencies_available():
    pytest.importorskip('openpyxl')
    from io import BytesIO
    from openpyxl import load_workbook
    from routers.excel import build_template_workbook

    wb = build_template_workbook()
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    loaded = load_workbook(stream, data_only=False)

    assert loaded.sheetnames == ['사용안내', '작성용', '검토용']
    assert loaded['작성용']['A2'].value == '연번'
    assert loaded['검토용']['E3'].value == '=IF(D3="","",LEN(D3))'
    assert loaded['작성용'].freeze_panes == 'A3'


def test_builder_contains_all_required_prompt_sections():
    import prompts.builder as builder

    assert '카. 대학명, 기관명, 기업명, 상호명, 강사명' in builder.COMMON_SYSTEM
    assert set(builder.AREA_PROMPTS) == {'자율활동', '동아리활동', '진로활동', '과목별 세부능력 및 특기사항', '행동특성 및 종합의견'}
    assert '임원 활동 시 입력된 직책과 재임기간만 정확히 반영' in builder.AREA_PROMPTS['자율활동']
    assert '심화탐구는 실제 근거 있을 때만 반영' in builder.AREA_PROMPTS['동아리활동']
    assert '진로검사는 결과 나열 금지' in builder.AREA_PROMPTS['진로활동']
    assert '인성·생활 태도보다 학업 역량·문제해결·자기주도 탐구 중심' in builder.AREA_PROMPTS['과목별 세부능력 및 특기사항']
    assert '430~500자 범위에서 작성하되 최대 300자 엄수' in builder.AREA_PROMPTS['행동특성 및 종합의견']
    assert '제외한항목' in builder.MODE_WRITE_SCHEMA
    assert '자가점검' in builder.MODE_WRITE_SCHEMA
    assert '위반사항' in builder.MODE_REVIEW_SCHEMA
    assert '수정안' in builder.MODE_REVIEW_SCHEMA
    assert '10자 이상 연속 일치' in builder.SIMILARITY_PROMPT
